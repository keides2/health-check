# Health Check
# Usage: $ Python3 health-check.py
# e.g:   $ python3 health-check.py
# Execution environment: Python3 on Windows10 or Linux

import os
import sys
# import requests     # pip install requests
import json
import re

import openpyxl as pyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
import datetime
import pprint as pp

# 保管場所
if os.name == 'nt':     # Windows
    FOLDER = "Z:\\path\\to\\health-check\\"
else:                   # Linux
    FOLDER = "/mnt/z/path/to/health-check/"

FILE_XLSX = "health-check.xlsx"
SHEETNAME = "12月"

# ファイル
FILE_EMAIL_ADDRESS = "email-address.txt"
FILE_EMAIL_BODY = "email-body.txt"
FILE_MENTION = "mention.xlsx"

# 日付から行番号取得用オフセット：20行目に、15日 があるのでオフセットは、4 (0 オリジン)
OFFSET = 4

# E-mail address 行
ROW_ADDRESS = 2
# チーム行
ROW_TM = 3

# C列 (0オリジン)
COLUMN_C = 2
# AI列 (0オリジン)
COLUMN_AI = 34

# 2次元タプルの各要素（Cellオブジェクト）から値を取得して2次元配列とする
def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])

# 1始まりの行番号・列番号で範囲を指定して2次元配列（リストのリスト）として取得
def get_list_2d(sheet, start_row, end_row, start_col, end_col):
    return get_value_list(sheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=start_col,
                                          max_col=end_col))

# エクセルのR1C1形式（行列番号）をアルファベット形式（アドレス）に変換する
def num2alpha(num):
    if num<=26:
        return chr(64+num)
    elif num%26==0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num%26)

# エクセルファイルにテーブル書式を挿入する
def insert_table(input_file, max_row):
    print(input_file, max_row)
    wb = load_workbook(input_file)
    ws = wb.active
    '''
    print(type(ws['A1']), ws.cell(row=1, column=1).value)
    print(type(ws['B1']), ws.cell(row=1, column=2).value)
    print(type(ws['A2']), ws.cell(row=2, column=1).value)
    print(type(ws['A3']), ws.cell(row=3, column=1).value)
    '''
    # 1-2行目削除、1行目に英語ヘッダを追加（日本語は、UserWarning: File may not be readable: column headings must be strings.）
    ws.delete_rows(1, 2)
    ws.insert_rows(1)
    ws['A1'] ="Team managers who have not filled in yet"

    # Add a default style with striped rows and banded columns
    mediumstyle = TableStyleInfo(name="TableStyleMedium9",
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=False)
    # create a table
    o = 'A1:A%d' % max_row  # "A1:B19"
    print(o)
    table = Table(displayName="Table1", ref=o, tableStyleInfo=mediumstyle)
    # table.tableStyleInfo = mediumstyle

    # add the table to the worksheet
    ws.add_table(table)

    # Saving the file
    wb.save(FOLDER + "mention_table.xlsx")
    wb.close()



# 開始
def main():
    # 今日の日付
    dt_now = datetime.datetime.now()
    print("Today: ", dt_now)

    # 月
    month = dt_now.strftime('%m月')
    print("This Month: ", month)

    # 日付
    # today = dt_now.strftime('%m/%d')

    # エクセルファイルのロード
    wb = load_workbook(filename=FOLDER+FILE_XLSX, read_only=True, data_only=True)

    # シートのロード
    sheet = wb[month]
    
    number_of_rows = sheet.max_row
    number_of_columns = sheet.max_column
    print("Rows: %d, Columns %d" % (number_of_rows, number_of_columns))

    # Store all values
    rows_iter = sheet.iter_rows(min_col=1, min_row=1, max_row=number_of_rows, max_col=number_of_columns)
    values = [[cell.value for cell in row] for row in rows_iter]
    
    # 本日のセル(B列?行)
    ROW_TODAY = dt_now.day + OFFSET
    # for test
    # ROW_TODAY = 15
    print("今日の行番号: " , ROW_TODAY+1)
    
    # Store today's value as generater
    # g = sheet.iter_rows(min_row=ROW_TODAY+1, max_row=ROW_TODAY+1, min_col=COLUMN_C, max_col=COLUMN_AI+1)
    # pp.pprint(list(g))

    # 本日の入力状況
    l_2d = get_value_list(sheet['C' + str(ROW_TODAY+1) + ':AH' + str(ROW_TODAY+1)])
    print("本日の入力状況")
    pp.pprint(l_2d, width=40)
    '''
    # start_row=3, end_row=4, start_col=3, end_col=5 = C3:E4
    l_2d = get_list_2d(sheet, 3, 4, 3, 5)
    pp.pprint(l_2d, width=40)

    # start_row=ROW_TODAY+1, end_row=ROW_TODAY+1, start_col=COLUMN_C, end_col=COLUMN_AI+1
    l_2d = get_list_2d(sheet, ROW_TODAY+1, ROW_TODAY+1, COLUMN_C, COLUMN_AI+1)
    pp.pprint(l_2d, width=40)
    '''

    # ロードしたExcelファイルを閉じる
    wb.close()

    # 未入力者通知用
    # メンション先のリスト（0オリジン）
    mention_list = []
    for i in range(COLUMN_C, COLUMN_AI+1, 2):
        # 有/無
        cell_symptoms1 = values[ROW_TODAY][i]
        print("i: %d, value: %s" % (i, cell_symptoms1))
        if (cell_symptoms1 is None):
            # 未入力の場合
            email_address = values[ROW_ADDRESS][i]
            print("Email_address: ", email_address)
            mention_list.append(email_address)
    
    # Teams - チャネル f26bxxxc.grp.o365.xxx.com@apac.teams.ms
    # mention_list.append('ad7xxxe0.grp.o365.xxx.com@apac.teams.ms')
    print(mention_list)

    # メンションリスト用テキストファイル作成・保存
    str_ = '\n'.join(mention_list)
    with open(FOLDER+FILE_EMAIL_ADDRESS, 'wt') as f:
        f.write(str_)
    
    # メンションリスト用エクセルファイルの作成・保存
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'health-check'

    sheet['A1'] = dt_now
    sheet['B1'] = '現在の入力状況です'
    sheet['A2'] = '以下のチームが未入力です'

    i = COLUMN_C + 1
    for m in mention_list:
        sheet['A' + str(i) ] = m
        i += 1

    wb.save(FOLDER+FILE_MENTION)
    wb.close()

    # テーブル挿入エクセルファイル "mention_table.xlsx" の作成 
    insert_table(FOLDER+FILE_MENTION, i-2)  # 先頭2行削除、1行追加分を追加で減算する

    
    # 管理者あて通知用メールの作成
    # メール本文
    mail_body = [[]]
    symptoms = []
    for i in range(COLUMN_C, COLUMN_AI+1, 2):
        # 有/無、詳細内容
        symptoms = []
        cell_TM = values[ROW_TM][i]
        cell_TM = cell_TM.replace('\n', '-')    # '\n' を '-' に置換
        symptoms.append(cell_TM)
        # 有/無
        cell_symptoms1 = values[ROW_TODAY][i]
        if (cell_symptoms1 is None):
            # 未入力の場合、None を文字に置き換え
            cell_symptoms1 = "▲未入力"
        
        symptoms.append(cell_symptoms1)

        # 詳細内容
        cell_symptoms2 = values[ROW_TODAY][i+1]
        if (cell_symptoms2 is None):
            # 未入力の場合、None を文字に置き換え
            cell_symptoms2 = "▲未入力"
        symptoms.append(cell_symptoms2)

        print("i: %d, TM: %s, Symptoms1: %s, Symtoms2: %s" % (i, cell_TM, cell_symptoms1, cell_symptoms2))
        
        mail_body.append(symptoms)

    print("mail_body: %s" % mail_body)
    print("\n")

    # メール本文用ファイル作成
    with open(FOLDER+FILE_EMAIL_BODY, 'w') as file:
        writer = csv.writer(file, delimiter='\t', lineterminator='\n')
        writer.writerows(mail_body)
    

    # 終了
    print("Done!")


if __name__ == '__main__':
    main()
    sys.exit(0)
